import argparse
import time
from pathlib import Path
from typing import Literal
import pandas as pd
import psutil
import openpyxl
import os
import tranco  # Ensure you import this if using it
from custom_command import LinkCountingCommand
from openwpm.command_sequence import CommandSequence
from openwpm.commands.browser_commands import GetCommand
from openwpm.config import BrowserParams, ManagerParams
from openwpm.storage.sql_provider import SQLiteStorageProvider
from openwpm.task_manager import TaskManager

# Define command-line arguments
parser = argparse.ArgumentParser()
parser.add_argument("--tranco", action="store_true", default=False)
parser.add_argument("--headless", action="store_true", default=False)
args = parser.parse_args()

# Read URLs from CSV and ensure proper formatting
sites = list(pd.read_csv("WebsiteLists/TargetList.csv")["url_domain"])
sites = [url if url.startswith(("http://", "https://")) else f"http://{url}" for url in sites]

if args.tranco:
    print("Loading tranco top sites list...")
    t = tranco.Tranco(cache=True, cache_dir=".tranco")
    latest_list = t.list()
    sites = ["http://" + x for x in latest_list.top(10)]

# Set display mode based on --headless argument
display_mode: Literal["native", "headless", "xvfb"] = "native"
if args.headless:
    display_mode = "headless"

# Configure browser and manager parameters
NUM_BROWSERS = 1
manager_params = ManagerParams(num_browsers=NUM_BROWSERS)
browser_params = [BrowserParams(display_mode=display_mode) for _ in range(NUM_BROWSERS)]

# Define the Excel file path
excel_file = Path("./datadir/ResourceUsage.xlsx")
# Get the name of the script
script_name = os.path.splitext(os.path.basename(__file__))[0]

# Check if the file exists
if not excel_file.exists():
    # Create a new workbook and sheet if file does not exist
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = "Sites"
    sheet['B1'] = "CPU Usage"
    sheet['C1'] = "Memory Usage"
    sheet['D1'] = "Disk Write"
    sheet['E1'] = "Disk Read"
    sheet['F1'] = "Disk Usage"
else:
    # Load the existing workbook and sheet
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

# Find the next empty row
next_row = sheet.max_row + 1

# Append the script name 
sheet.append([script_name])

# Save and close the workbook
workbook.save(excel_file)

for browser_param in browser_params:
    # Record HTTP Requests and Responses
    browser_param.http_instrument = True
    # Record cookie changes
    browser_param.cookie_instrument = True
    # Record Navigations
    browser_param.navigation_instrument = True
    # Record JS Web API calls
    browser_param.js_instrument = True
    # Record DNS resolution
    browser_param.dns_instrument = True
    # Performs action to prevent the platform from being detected as a bot
    browser_param.bot_mitigation = True
    # Accepts all third party cookies
    browser_param.tp_cookies = "always"
    # Turn off donottrack in browser
    browser_param.donottrack = False
    # Load a browser profile
    browser_param.seed_tar = Path("./ProfileLists/ProfileFN/profile.tar.gz")
    browser_param.maximum_profile_size = 50 * (10**20)  # 50 MB

manager_params.data_directory = Path("./datadir/")
manager_params.log_path = Path("./datadir/openwpm.log")

# Initialize and run the TaskManager
with TaskManager(
    manager_params,
    browser_params,
    SQLiteStorageProvider(Path("./datadir/crawl-dataFNSL.sqlite")),
    None,
) as manager:
    for index, site in enumerate(sites):
        def callback(success: bool, val: str = site) -> None:
            print(f"CommandSequence for {val} ran {'successfully' if success else 'unsuccessfully'}")

        # Create a command sequence for each site
        command_sequence = CommandSequence(
            site,
            reset=True, # Enables a stateless crawl
            site_rank=index,
            callback=callback,
        )

        # Start by visiting the page
        
        command_sequence.append_command(GetCommand(url=site, sleep=3), timeout=60)
        # Have a look at custom_command.py to see how to implement your own command
        command_sequence.append_command(LinkCountingCommand())

        # Run commands across all browsers (simple parallelization)
        manager.execute_command_sequence(command_sequence)

